#!/usr/bin/python 
#coding:utf-8
"""
This program will convert the input picture file to an Excel file. Each pixel's color will 
be represented as the same background color of the corresponding cell in the Excel file. It only supports jpg image. 

Usage:
    PictureToExcelCellColorConverter.py input.jpg
 
@author: Haiyang Cui
"""

import sys
from pathlib import Path 
from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import math
import os

def main(argv):
    if len(sys.argv) != 2:
        sys.exit("Usage: PictureToExcelCellColorConverter.py input.jpg") 
    inputfile = argv[1]   
    m_file = Path(inputfile)
    if not m_file.exists():
        sys.exit("The input file does not exist.")
        
    print( 'The input picture file is '+ inputfile)
    
    pictureToExcel(inputfile)
 
 
def pictureToPixel(picFile):
    im = Image.open(picFile)  
    pixels = im.load()    
    return(im.size, pixels)

def pictureToExcel(picFile):
    maxWidth  = 200
    maxHeight = 200
    newFileName = resizePicture(picFile,maxWidth, maxHeight)
    size,pixels = pictureToPixel(newFileName)
    (prefix, sep, suffix) = picFile.rpartition('.')  
    excelFile = prefix + '.xlsx'
    print('Will create file '+ excelFile)
    
    wb = openpyxl.Workbook()
    ws = wb.active

    ######################
    for rowIndex in range(1,size[1]):
        ws.row_dimensions[rowIndex].height = 10
        for colIndex in range(1,size[0]):
            #print(get_column_letter(colIndex))
            ws.column_dimensions[get_column_letter(colIndex)].width = 2.5
            r,g,b=pixels[colIndex,rowIndex]
            hexColor= rgb_to_hex(r,g,b)
            redFill = PatternFill(start_color=hexColor[1:],end_color=hexColor[1:], fill_type='solid')
            ws.cell(row=rowIndex,column=colIndex).fill = redFill   
    ######################        
    wb.save(excelFile)

def rgb_to_hex(red, green, blue):
    """Return color as #rrggbb for the given color values."""
    return '#%02x%02x%02x' % (red, green, blue)

def resizePicture(picFile,maxWidth, maxHeight):
    im = Image.open(picFile)  
    ratio = min( min(maxWidth,im.size[0])/im.size[0], min(maxHeight,im.size[1])/im.size[1] )
    newWidth = math.floor(im.size[0]*ratio)
    nweHeight = math.floor(im.size[1]*ratio)
    im.thumbnail((newWidth,nweHeight),Image.ANTIALIAS)
    dir_name, file_name = os.path.split(picFile)
    (prefix, sep, suffix) = file_name.rpartition('.')  
    
    newFileName = os.path.join(dir_name, "ResizedPicture"+sep+suffix)
    im.save(newFileName)
    return(newFileName)
    
if __name__ == "__main__":
   main(sys.argv)
   
   